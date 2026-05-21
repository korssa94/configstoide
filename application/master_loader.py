import openpyxl


def load_master_map(master_file):
    """Читает Мастер-конфигуратор и строит карту {имя_конфигуратора: [список_ПЛК]}.

    Структура мастер-файла: первый столбец — категория ("Контроллер", "Входы/выходы",
    "Сигнализации" и т.д.), остальные столбцы — данные по каждому ПЛК. В строке
    "Контроллер" перечислены имена ПЛК, в остальных — имена связанных файлов.
    """
    wb_master = openpyxl.load_workbook(master_file, data_only=True)
    ws_m = wb_master.active
    master_map = {}
    plc_headers = []
    plc_col_indices = []

    for row in ws_m.iter_rows(values_only=True):
        if row and str(row[0]).strip() == "Контроллер":
            plc_headers = row
            plc_col_indices = [i for i, val in enumerate(row) if i > 0 and val]
            break

    for row in ws_m.iter_rows(values_only=True):
        if not row or str(row[0]).strip() == "Контроллер":
            continue
        for idx in plc_col_indices:
            raw_val = row[idx]
            if raw_val and str(raw_val).strip() != "None":
                f_name = str(raw_val).strip()
                p_name = str(plc_headers[idx]).strip()
                if f_name not in master_map:
                    master_map[f_name] = []
                if p_name not in master_map[f_name]:
                    master_map[f_name].append(p_name)

    return master_map