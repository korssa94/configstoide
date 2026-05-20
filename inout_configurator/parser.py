# parsers/te5_parser.py
import openpyxl
import os
import datetime
import re
from shared.parsers.base_parser import BaseParser
import inout_configurator.models as models

class TE5Parser(BaseParser):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
    
    @property
    def all_parsed_objects(self):
        """Свойство, возвращающее все распарсенные объекты одним списком.
        Необходимо для совместимости с app.py и универсальными утилитами."""
        all_objs = []
        for obj_list in self.objects.values():
            all_objs.extend(obj_list)
        return all_objs
    
    def parse(self, clean_name, active_rules, force=False):
        try:
            wb_data = openpyxl.load_workbook(self.filepath, data_only=True)
            self.file_author = wb_data.properties.lastModifiedBy or "Unknown"
            if wb_data.properties.modified:
                utc_time = wb_data.properties.modified
                local_time = utc_time.replace(tzinfo=datetime.timezone.utc).astimezone(None)
                self.file_save_time = local_time.strftime("%Y-%m-%d %H:%M:%S")
            
            algo_folder = self.config.DEFAULT_ALGO_FOLDER
            valid_subsystems = []
            
            if self.config.SHEET_SETTINGS in wb_data.sheetnames:
                reading_subs = False
                for row in wb_data[self.config.SHEET_SETTINGS].iter_rows(values_only=True):
                    first_cell = str(row[0]).strip() if row and row[0] else ""
                    if first_cell == "Папка с алгоритмами":
                        algo_folder = str(row[1]).strip() if row[1] else algo_folder
                    elif first_cell == "Подсистемы":
                        reading_subs = True
                        continue
                    if reading_subs and row and (row[0] or row[1]):
                        name, code, num = (str(row[0]).strip() if row[0] else "", 
                                           str(row[1]).strip() if row[1] else "", 
                                           str(row[2]).strip() if row[2] else "")
                        parts = [p for p in [name, code, num] if p and p != "None"]
                        valid_subsystems.append(" - ".join(parts))

            processed_sheets = set()
            all_stats = {} # <-- Создаем словарь для статистики

            for model_name, settings in self.config.MODEL_SETTINGS.items():
                sheet_name = settings["sheet_name"]
                
                self.objects[model_name] = []
                
                if sheet_name in wb_data.sheetnames:
                    processed_sheets.add(sheet_name)
                    
                    ws_data = wb_data[sheet_name]
                    col_map = self.get_column_mapping(ws_data, self.config.HEADER_ROW)
                    
                    ModelClass = getattr(models, model_name, None)
                    if ModelClass:
                        stats = self._parse_sheet(ws_data, col_map, clean_name, active_rules, ModelClass, valid_subsystems, force)
                        # <-- Просто сохраняем статистику, пока ничего не печатаем
                        all_stats[model_name] = {
                            "sheet_name": sheet_name, 
                            "created": stats["created"], 
                            "skipped": stats["skipped"],
                            "exclusions": stats["exclusions"],
                            "errors": stats["errors"] # <-- добавили
                        }
                    else:
                        self.log(f'Класс модели {model_name} не найден в коде!', level="ERROR")

            # --- ПЕРЕКРЕСТНЫЕ ПРОВЕРКИ ---
            self._cross_validate(force, clean_name, all_stats)

            # 1. ВЫВОД ЛОГОВ АНАЛИЗА И ОШИБОК (СГРУППИРОВАНО ПО ЛИСТАМ)
            if not force:
                for m_name, settings in self.config.MODEL_SETTINGS.items():
                    if m_name in all_stats:
                        sht_name = all_stats[m_name]["sheet_name"]
                        # Сначала печатаем заголовок анализа листа
                        self.log(f'Анализ листа "{sht_name}".')
                        
                        # Если на листе есть ошибки (свои или перекрестные), выводим их сразу под заголовком
                        if all_stats[m_name]["errors"]:
                            for err_msg in all_stats[m_name]["errors"]:
                                self.log(err_msg, level="ERROR")
                                
                # Проверяем, есть ли листы, которые парсер видит, но не обрабатывает
                all_known_sheets = [self.config.SHEET_AI, self.config.SHEET_AO, self.config.SHEET_DI, self.config.SHEET_DO, self.config.SHEET_FPL]
                for other in all_known_sheets:
                    if other in wb_data.sheetnames and other not in processed_sheets:
                        self.log(f'Лист "{other}" найден, но не обрабатывается текущей версией.', level="INFO")

            # 2. И ТОЛЬКО ПОТОМ ОСТАНАВЛИВАЕМ ПРОЦЕСС
            if self.errors and not force:
                self.log("Ожидание решения пользователя...", level="WARNING")
                return False

            # --- ВЫВОД ИТОГОВ ГЕНЕРАЦИИ ---
            for model_name, settings in self.config.MODEL_SETTINGS.items():
                if model_name in all_stats:
                    s_name = all_stats[model_name]["sheet_name"]
                    c = all_stats[model_name]["created"]
                    s = all_stats[model_name]["skipped"]
                    excl_list = all_stats[model_name]["exclusions"]

                    # 1. Печатаем итоги по листу
                    msg = f'Лист "{s_name}": создано {c} сигналов'
                    if s > 0:
                        msg += f', пропущено {s} из-за ошибок в конфигурации'
                    self.log(f"{msg}.")

                    # 2. Сразу под итогами выводим список "выживших из ума" параметров :)
                    if force and excl_list:
                        for warn_msg in excl_list:
                            self.log(warn_msg, level="WARNING")

            # Генерируем subsystem.gvl для всех контроллеров (один раз, а не на каждой модели)
            self.generate_subsystems_gvl(valid_subsystems, "входов/выходов")

            self._generate_files(algo_folder)
            return True
        except Exception as e:
            self.log(f"Ошибка в ТЭ5 парсере: {str(e)}", level="ERROR")
            return False

    def _parse_sheet(self, ws_data, col_map, clean_name, active_rules, ModelClass, valid_subsystems, force):
        idx_create = self.find_col_idx(col_map, self.config.COL_CREATE_CODE)
        counts = {"created": 0, "skipped": 0, "exclusions": [], "errors": []}
        seen_addresses = {}

        # Ищем индекс колонки "Номер"
        idx_num = self.find_col_idx(col_map, getattr(self.config, 'COL_NUM', '№'))

        for row in range(self.config.DATA_START_ROW, ws_data.max_row + 1):
            val_raw = ws_data.cell(row=row, column=idx_create).value if idx_create else ""
            
            # ЖЕЛЕЗОБЕТОННАЯ ЗАЩИТА: Если номер пустой — это конец таблицы, выходим из цикла!
            num_val = ws_data.cell(row=row, column=idx_num).value if idx_num else ""
            if num_val is None or str(num_val).strip() == "":
                break 
            
            # --- ФОРМИРОВАНИЕ КАРТЫ ПОКРАСКИ НА СТОРОНЕ ПАРСЕРА ---
            sheet_name = ws_data.title
            if not hasattr(self, 'rows_to_color'):
                self.rows_to_color = {}
            if sheet_name not in self.rows_to_color:
                self.rows_to_color[sheet_name] = {}
                
            # Для ТЭ5 резервом считается всё, где "Создавать код" не равен 1
            is_reserve = str(val_raw).strip() != "1"
            self.rows_to_color[sheet_name][row] = is_reserve
            
            if is_reserve: continue
            
            def get_val(key):
                idx = self.find_col_idx(col_map, key)
                return ws_data.cell(row=row, column=idx).value if idx else ""

            # Получаем индекс базового столбца "Токовые уставки" для Шлейфов
            idx_sp = self.find_col_idx(col_map, getattr(self.config, 'COL_SP_BASE', 'Токовые уставки'))
            idx_dc = self.find_col_idx(col_map, self.config.COL_DEVICE_CLAMP)

            signal_data = {
                'alg_name': get_val(self.config.COL_ALG_NAME),
                'desc': get_val(self.config.COL_DESC),
                'tech_name': get_val(self.config.COL_TECH_NAME),
                'short_name': get_val(self.config.COL_SHORT_NAME),
                'units': get_val(self.config.COL_UNITS),
                'electrical_units': get_val(self.config.COL_ELEC_UNITS),
                'tag_prefix': get_val(self.config.COL_TAG_PREFIX),
                'device': get_val(self.config.COL_DEVICE),
                'crate': get_val(self.config.COL_CRATE),
                'module': get_val(self.config.COL_MODULE),
                'channel': get_val(self.config.COL_CHANNEL),
                'module_type': get_val(self.config.COL_MODULE_TYPE),
                'signal_type': get_val(self.config.COL_SIGNAL_TYPE),
                'signal_char': get_val(self.config.COL_SIGNAL_CHAR),
                'subsystem': get_val(self.config.COL_SUBSYSTEM),
                'server_cycle': get_val(self.config.COL_SERVER_CYCLE),
                'type': get_val(self.config.COL_TYPE),
                'circuit_control': get_val(self.config.COL_CIRCUIT_CONTROL),
                'min_val': get_val(self.config.COL_MIN_VAL),
                'max_val': get_val(self.config.COL_MAX_VAL),
                'll': get_val(self.config.COL_LL),
                'l1': get_val(self.config.COL_L1),
                'l': get_val(self.config.COL_L),
                'h': get_val(self.config.COL_H),
                'h1': get_val(self.config.COL_H1),
                'hh': get_val(self.config.COL_HH),
                'hysteresis': get_val(self.config.COL_HYSTERESIS),
                'max_rate': get_val(self.config.COL_MAX_RATE),
                'k': get_val(self.config.COL_FREQ_COEF),
                'digit': get_val(self.config.COL_PRECISION),
                'group': get_val(self.config.COL_GROUPS),
                'device_clamp': str(ws_data.cell(row=row, column=idx_dc).value).replace('\n', ' ').strip() if idx_dc and ws_data.cell(row=row, column=idx_dc).value is not None else "",
                'clamp': str(ws_data.cell(row=row, column=idx_dc + 1).value).replace('\n', ' ').strip() if idx_dc and ws_data.cell(row=row, column=idx_dc + 1).value is not None else "",
                
                # --- Новые поля для Шлейфов (Tfpl) ---
                'f_type': get_val(getattr(self.config, 'COL_F_TYPE', '')),
                'voting': get_val(getattr(self.config, 'COL_VOTING', '')),
                'sp0': ws_data.cell(row=row, column=idx_sp).value if idx_sp else "",
                'sp1': ws_data.cell(row=row, column=idx_sp + 1).value if idx_sp else "",
                'sp2': ws_data.cell(row=row, column=idx_sp + 2).value if idx_sp else "",
                'sp3': ws_data.cell(row=row, column=idx_sp + 3).value if idx_sp else "",
                'sp4': ws_data.cell(row=row, column=idx_sp + 4).value if idx_sp else ""
            }
            
            obj = ModelClass(row_number=row, **signal_data)
            type_rules = active_rules.get(ModelClass.__name__, {})
            errors = obj.validate(type_rules, valid_subsystems) 
            
            if type_rules.get("duplicate_address", True):
                addr_key = obj.get_full_address()
                if addr_key and addr_key in seen_addresses:
                    errors.append({"msg": f"Дубликат адреса: строка {seen_addresses[addr_key]}", "field": "address"})
                elif addr_key: seen_addresses[addr_key] = row

            if errors:
                counts["skipped"] += 1 # Считаем пропуски
                for err in errors:
                    self.errors.append({"Файл": clean_name, "Строка": row, "Имя": obj.alg_name or "---", "Ошибка": err["msg"]})
                    if not force:
                        # Убрали мгновенный self.log, копим ошибки
                        counts["errors"].append(f"Строка {row} [{obj.alg_name or '---'}]: {err['msg']}")
                
                # При принудительной генерации сохраняем WARNING в список для вывода позже
                if force:
                    counts["exclusions"].append(f"Переменная со строки {row}, наименованием '{obj.description or '---'}', алг. именем '{obj.alg_name or '---'}' исключена из генерации, т.к. неправильно сконфигурирована")
            else:
                self.objects[ModelClass.__name__].append(obj)
                counts["created"] += 1
        return counts

    def _generate_files(self, algo_folder):
        fn = os.path.basename(self.filepath)
        excel_v = (re.search(r'v\d+\.\d+\.\d+', fn).group(0) if re.search(r'v\d+\.\d+\.\d+', fn) else "none")
        
        # --- Инициализация файла сброса (reset_io_var_stat.st) ---
        h_reset = self.generate_header(algo_folder, self.file_author, self.file_save_time, excel_v, "Используется для сброса всех VAR_STAT входов/выходов")
        reset_txt = h_reset + "FUNCTION reset_io_var_stat : BOOL\nVAR_INPUT\nEND_VAR\n"
        reset_txt += "//++++++< Content:Implementation >++++++++++++++++++++++++++++++++++++++++++++//\n"
        has_reset_content = False
        
        for model_name, obj_list in self.objects.items():
            if not obj_list: continue
            
            settings = self.config.MODEL_SETTINGS[model_name]
            global_var = settings["global_var"] 
            needs_cycle = settings.get("needs_cycle", False)
            plc_common_type = settings.get("plc_common_type", model_name)
            
            # --- 1. Добавляем логику сброса для текущей модели ---
            has_reset_content = True
            reset_txt += f"\n//Сбрасываем для {global_var}\n"
            reset_txt += f"{global_var}.common.flt_aggregated := FALSE;\n"
            reset_txt += f"{global_var}.common.sim_aggregated := FALSE;\n"
            if model_name == "Taipar":
                reset_txt += f"{global_var}.common.ucv_active := FALSE;\n"

            # --- 2. Генерация GVL ---
            h_gvl = self.generate_header(algo_folder, self.file_author, self.file_save_time, excel_v, settings["desc_gvl"])
            gvl_txt = h_gvl + f"{{attribute 'qualified_only'}}\n"
            gvl_txt += f"{{attribute 'symbol' := 'readwrite'}}\n"
            gvl_txt += f"VAR_GLOBAL\n"
            gvl_txt += f"\t{{attribute 'symbol' := 'none'}}\n"
            gvl_txt += f"\tcommon : {plc_common_type};\n"
            gvl_txt += "\n".join([o.get_gvl_string() for o in obj_list]) + "\nEND_VAR\n"

            # --- 3. Генерация ST ---
            h_st = self.generate_header(algo_folder, self.file_author, self.file_save_time, excel_v, settings["desc_st"])
            
            if needs_cycle:
                st_txt = h_st + f"FUNCTION {global_var}_update : BOOL\nVAR_INPUT\n\tcycle : REAL;\nEND_VAR\n"
            else:
                st_txt = h_st + f"FUNCTION {global_var}_update : BOOL\n"
                
            st_txt += "//++++++< Content:Implementation >++++++++++++++++++++++++++++++++++++++++++++//\n"
            
            if needs_cycle:
                st_txt += f"{global_var}.common.cycle := cycle;\n\n"
                
            st_txt += "\n".join([o.get_st_string() for o in obj_list]) + "\n"

            # --- 4. Сохранение файлов моделей в список ---
            for ctrl in self.matched_ctrls:
                tdir = os.path.join(self.base_dir, self.config.SOURCE_FOLDER, ctrl, algo_folder)
                self.files_to_write.append({"path": os.path.join(tdir, settings["file_gvl"]), "text": gvl_txt})
                self.files_to_write.append({"path": os.path.join(tdir, settings["file_st"]), "text": st_txt})

        # --- 5. Сохранение файла сброса в список (один на папку алгоритмов) ---
        if has_reset_content:
            for ctrl in self.matched_ctrls:
                tdir = os.path.join(self.base_dir, self.config.SOURCE_FOLDER, ctrl, algo_folder)
                self.files_to_write.append({"path": os.path.join(tdir, "reset_io_var_stat.st"), "text": reset_txt})
    
    def _cross_validate(self, force, clean_name, all_stats):
        """Выполняет перекрестные проверки между уже собранными листами.

        Каждый объект проверяется одним правилом — определяемым по значению
        circuit_control. Правила описаны декларативной таблицей CROSS_CHECKS.
        """
        lookups = {
            "taipar_names": {obj.alg_name for obj in self.objects.get("Taipar", [])},
            "tdipar_names": {obj.alg_name for obj in self.objects.get("Tdipar", [])},
        }

        # Декларативные правила кросс-проверки:
        # модель → список (триггер_circuit_control, префикс_имени_контроля, lookup, лист_для_сообщения)
        cross_checks_by_model = {
            "Taopar": [
                ("1", "kcdi_", "tdipar_names", "Вх.Д сигн."),
                ("2", "kcao_", "taipar_names", "Вх.А сигн."),
            ],
            "Tdipar": [
                ("3", "kcdi_", "tdipar_names", "Вх.Д сигн."),
            ],
            "Tdopar": [
                ("1", "kcdo_", "tdipar_names", "Вх.Д сигн."),
                ("2", "kcdo_", "taipar_names", "Вх.А сигн."),
            ],
        }

        for model_name, rules in cross_checks_by_model.items():
            if model_name not in self.objects:
                continue

            valid_objects = []
            for obj in self.objects[model_name]:
                is_valid = True
                cc = str(obj.circuit_control).strip()

                # На объект может сработать ровно одно правило — триггеры взаимоисключающие
                for trigger, prefix, lookup_key, sheet_label in rules:
                    if cc != trigger:
                        continue
                    expected_name = f"{prefix}{obj.alg_name}"
                    if expected_name not in lookups[lookup_key]:
                        msg = f"Отсутствует параметр контроля цепи '{expected_name}' на листе {sheet_label}."
                        self._record_cross_validation_error(
                            model_name, obj, msg, clean_name, force, all_stats
                        )
                        is_valid = False
                    break

                if is_valid or not force:
                    valid_objects.append(obj)

            # При force-режиме отсеиваем бракованные объекты из self.objects
            if force:
                self.objects[model_name] = valid_objects

    def _record_cross_validation_error(self, model_name, obj, msg, clean_name, force, all_stats):
        """Регистрирует ошибку кросс-проверки: в self.errors, в логе листа, корректирует статистику."""
        self.errors.append({
            "Файл": clean_name,
            "Строка": obj.row_number,
            "Имя": obj.alg_name,
            "Ошибка": f"Кросс-чек: {msg}"
        })

        if model_name not in all_stats:
            return

        if not force:
            all_stats[model_name]["errors"].append(
                f"Строка {obj.row_number} [{obj.alg_name or '---'}]: Перекрестная проверка: {msg}"
            )
        else:
            all_stats[model_name]["exclusions"].append(
                f"Переменная со строки {obj.row_number}, наименованием '{obj.description or '---'}', "
                f"алг. именем '{obj.alg_name or '---'}' исключена из генерации (провал перекрестной проверки)"
            )

        all_stats[model_name]["created"] -= 1
        all_stats[model_name]["skipped"] += 1
