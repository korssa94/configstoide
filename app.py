# app.py
import sys
import os

# --- БЛОК ОЧИСТКИ КЭША ---
# Добавили "application" в список очистки, чтобы Streamlit видел изменения в наших новых файлах
project_modules = [m for m in sys.modules if any(k in m for k in ["parsers", "models", "settings", "application"])]
for module in project_modules:
    del sys.modules[module]

import streamlit as st
import openpyxl
import re
import urllib.parse
from settings import AppConfig, TE5Config, AlarmConfig
from parsers import TE5Parser, AlarmParser
from documentation.document_updater import update_configurator_document
from documentation.text_updater import update_configurator_texts
from documentation.color_updater import update_configurator_colors

# Импорты из нашей новой архитектуры
from application.scanner import find_plcopen_xmls
from application.settings_manager import load_settings, save_settings, sync_addon_path, on_checkbox_change
from application.logger import add_log, render_logs

CONFIG_REGISTRY = {
    "TE5": {"config": TE5Config, "parser": TE5Parser},
    "Alarms": {"config": AlarmConfig, "parser": AlarmParser}
}

# Выполняем синхронизацию только 1 раз при старте приложения
if "addon_synced" not in st.session_state:
    sync_addon_path()
    st.session_state.addon_synced = True

st.set_page_config(page_title=f"PLC Generator Pro Max v{AppConfig.APP_VERSION}", layout="wide")
st.title(f"🏭 PLC Code Generator Pro Max GT Edition v{AppConfig.APP_VERSION}")

# --- КАСТОМНЫЙ CSS ДЛЯ STREAMLIT ---
st.markdown("""
<style>
    /* 1. Контейнер: разрешаем перенос пузырьков на новые строки */
    .stMultiSelect div[data-baseweb="select"] > div:first-child {
        flex-wrap: wrap !important;
        padding-bottom: 4px !important;
    }
    
    /* 2. Пузырьки: снимаем жесткое ограничение по ширине, но не даем вылезти за пределы поля */
    .stMultiSelect [data-baseweb="tag"] {
        max-width: 100% !important; 
        margin-bottom: 4px !important;
        margin-top: 4px !important;
    }
    
    /* 3. Текст внутри пузырька: возвращаем СТРОГО в одну строку */
    .stMultiSelect [data-baseweb="tag"] span {
        white-space: nowrap !important; /* Текст в одну строку */
        overflow: hidden !important;
        text-overflow: ellipsis !important; /* Троеточие только если не влезает в экран */
        max-width: none !important;
    }
</style>
""", unsafe_allow_html=True)

user_settings = load_settings()

# Инициализация состояний
if 'logs' not in st.session_state: st.session_state.logs = []
if 'analyzed' not in st.session_state: st.session_state.analyzed = False
if 'process_step' not in st.session_state: st.session_state.process_step = None
if 'ms_counter' not in st.session_state: st.session_state.ms_counter = 0
if 'current_selection' not in st.session_state: st.session_state.current_selection = []
if 'selection_initialized' not in st.session_state: st.session_state.selection_initialized = False

target_file_from_excel = st.query_params.get("target_file")
skip_paint_global = bool(target_file_from_excel)

if target_file_from_excel:
    target_file_from_excel = urllib.parse.unquote(target_file_from_excel).strip('"')
    current_file_dir = os.path.normpath(os.path.dirname(target_file_from_excel))
else:
    new_dir = st.text_input("📁 Путь к папке Configs:", value=user_settings.get("base_dir", ""))
    if new_dir != user_settings.get("base_dir", ""):
        user_settings["base_dir"] = new_dir
        save_settings(user_settings)
        st.rerun()
    current_file_dir = new_dir

if current_file_dir and os.path.exists(current_file_dir):
    target_files = []
    master_file = None
    project_root = None 
    check_dir = current_file_dir
    
    while True:
        folder_name = os.path.basename(check_dir).lower()
        potential_masters = [os.path.join(check_dir, f) for f in os.listdir(check_dir) if AppConfig.KEYWORD_MASTER in f and f.endswith(('.xlsx', '.xlsm')) and not f.startswith('~$')]
        if potential_masters and master_file is None: master_file = potential_masters[0]
        if folder_name in ["config", "configs"]:
            project_root = check_dir
            break
        parent = os.path.dirname(check_dir)
        if parent == check_dir: break
        check_dir = parent
        
    final_root = project_root if project_root else (os.path.dirname(master_file) if master_file else current_file_dir)

    active_keywords = []
    for v in CONFIG_REGISTRY.values():
        kw = v["config"].KEYWORD_FILE
        if isinstance(kw, list):
            active_keywords.extend(kw)
        else:
            active_keywords.append(kw)

    for root, dirs, files in os.walk(current_file_dir):
        for f in files:
            if f.startswith('~$') or not f.endswith(('.xlsx', '.xlsm')): continue
            if any(k in f for k in active_keywords):
                target_files.append(os.path.join(root, f))

    if not master_file:
        st.error(f"❌ Мастер-конфигуратор не найден!")
    else:
        default_sel = [f for f in target_files if target_file_from_excel and os.path.normpath(f) == os.path.normpath(target_file_from_excel)]
        
        if not st.session_state.selection_initialized:
            st.session_state.current_selection = default_sel
            st.session_state.selection_initialized = True
            
        def on_ms_change():
            current_counter = st.session_state.get("ms_counter", 0)
            widget_key = f"ms_{current_counter}"
            if widget_key in st.session_state:
                st.session_state.current_selection = st.session_state[widget_key]
            st.session_state.ms_counter = current_counter + 1

        selected_files = st.multiselect(
            "📝 Конфигураторы:", 
            options=target_files, 
            format_func=lambda x: os.path.basename(x), 
            default=st.session_state.current_selection,
            key=f"ms_{st.session_state.ms_counter}",
            on_change=on_ms_change
        )
        
        # --- БЛОК ОПЦИЙ И ВЫБОРА XML ---
        st.subheader("🛠 Настройки процесса")
        col_opt1, col_opt2 = st.columns(2)

        with col_opt1:
            do_sources = st.checkbox("Создать исходники", value=True)
            do_coloring = st.checkbox("Обновить цвета ячеек", 
                          value=user_settings.get("do_coloring", False), 
                          key="do_coloring", 
                          on_change=on_checkbox_change, 
                          args=("do_coloring",))

        with col_opt2:
            do_translation = st.checkbox("Обновить текстовое описание", 
                             value=user_settings.get("do_translation", True), 
                             key="do_translation", 
                             on_change=on_checkbox_change, 
                             args=("do_translation",))
            do_document = st.checkbox("Обновить документ", 
                          value=user_settings.get("do_document", True), 
                          key="do_document", 
                          on_change=on_checkbox_change, 
                          args=("do_document",))

        selected_xml_path = None
        if do_translation and final_root:
            found_xmls, debug_info = find_plcopen_xmls(final_root)
            if found_xmls:
                xml_options = {f"[{f['module']}] {f['name']} ({f['size']//1024} KB)": f['path'] for f in found_xmls}
                selected_label = st.selectbox(
                    "Выберите файл проекта (PLCopen XML):", 
                    options=list(xml_options.keys()),
                    index=0
                )
                selected_xml_path = xml_options[selected_label]
            else:
                st.warning(f"⚠️ Файлы .plcopen.xml не найдены в подпапках {AppConfig.SOURCE_EXPORT_FOLDER}")
                with st.expander("🛠 Дебаг поиска"):
                    for line in debug_info: st.text(line)
                do_translation = False

        # --- КНОПКА ЗАПУСКА ---
        if st.button("🔍 Запустить процесс", type="primary") or (target_file_from_excel and not st.session_state.analyzed):
            st.session_state.process_step = "analyze"
            st.session_state.analyzed = True
            
        # --- БЛОК ОЖИДАНИЯ РЕШЕНИЯ ---
        if st.session_state.process_step == "awaiting_confirm":
            st.error(f"❌ Найдено ошибок: {len(st.session_state.file_errors)}. Генерация приостановлена!")
            col_btn1, col_btn2 = st.columns(2) 
            with col_btn1:
                if st.button("⚠️ Всё равно создать (без ошибочных)", type="primary", use_container_width=True):
                    st.session_state.process_step = "generate_forced"
                    st.rerun()
            with col_btn2:
                if st.button("🛑 Отменить генерацию", use_container_width=True):
                    st.session_state.process_step = "done"
                    add_log("🛑 Генерация отменена пользователем.", level="ERROR")
                    st.rerun()
        
        # --- ОСНОВНАЯ ЛОГИКА ---
        if st.session_state.process_step in ["analyze", "generate_forced"]:
            force = (st.session_state.process_step == "generate_forced")
            
            if not force:
                st.session_state.logs = []
                st.session_state.failed_rows_cache = {} # Очищаем кэш ошибочных строк перед каждым новым анализом
                add_log("🚀 Старт процесса генерации")
            else:
                add_log("🚀 Принудительная генерация (ошибочные сигналы исключены)")
                
            st.session_state.file_errors = []
            st.session_state.files_to_write = []
            
            wb_master = openpyxl.load_workbook(master_file, data_only=True)
            ws_m = wb_master.active
            master_map = {}
            plc_headers = []
            plc_col_indices = []

            for row_idx, row in enumerate(ws_m.iter_rows(values_only=True)):
                if row and str(row[0]).strip() == "Контроллер":
                    plc_headers = row
                    plc_col_indices = [i for i, val in enumerate(row) if i > 0 and val]
                    break

            for row in ws_m.iter_rows(values_only=True):
                if not row or str(row[0]).strip() == "Контроллер": continue
                for idx in plc_col_indices:
                    raw_val = row[idx]
                    if raw_val and str(raw_val).strip() != "None":
                        f_name = str(raw_val).strip()
                        p_name = str(plc_headers[idx]).strip()
                        if f_name not in master_map: master_map[f_name] = []
                        if p_name not in master_map[f_name]: master_map[f_name].append(p_name)

            for fp in selected_files:
                file_name = os.path.basename(fp)
                clean_n = re.sub(r'\s+v\d+\.\d+\.\d+.*$', '', os.path.splitext(file_name)[0]).strip()
                file_type = next((k for k, v in CONFIG_REGISTRY.items() if (any(kw in file_name for kw in v["config"].KEYWORD_FILE) if isinstance(v["config"].KEYWORD_FILE, list) else v["config"].KEYWORD_FILE in file_name)), None)
                if not file_type: continue
                
                ctrls = master_map.get(clean_n, [])
                if not force: add_log(f"📂 Анализ: {clean_n} (ПЛК: {len(ctrls)})")
                
                if ctrls:
                    parser_class = CONFIG_REGISTRY[file_type]["parser"]
                    config_class = CONFIG_REGISTRY[file_type]["config"]
                    parser = parser_class(fp, final_root, ctrls, config_class, logger=add_log)
                    
                    # Прокидываем путь к Мастеру внутрь парсера для кросс-проверок
                    parser.master_file = master_file 
                    
                    is_ok = parser.parse(clean_n, user_settings.get("validations", {}).get(file_type, {}), force=force)
                    
                    # --- БЛОК ОБНОВЛЕНИЯ ТЕКСТОВ ---
                    if (is_ok or force) and do_translation and selected_xml_path and getattr(config_class, 'SUPPORT_TEXT_UPDATE', False):
                        from parsers.xml_parser import build_xml_cache
                        from documentation.condition_translator import translate_condition
                        
                        add_log(f"📝 Перевод кода в текст с использованием XML-кэша...")
                        xml_cache = build_xml_cache(selected_xml_path, target_configs=ctrls, logger=add_log)
                        st.session_state.xml_cache = xml_cache
                        
                        parsed_objects = getattr(parser, 'all_parsed_objects', [])
                        
                        for obj in parsed_objects:
                            if hasattr(obj, 'trigger_cond') and obj.trigger_cond:
                                obj.trigger_text = translate_condition(obj.trigger_cond, xml_cache)
                            if hasattr(obj, 'fault_cond') and obj.fault_cond:
                                obj.fault_text = translate_condition(obj.fault_cond, xml_cache)
                            if hasattr(obj, 'set_code') and obj.set_code:
                                obj.set_text = translate_condition(obj.set_code, xml_cache)
                            if hasattr(obj, 'reset_code') and obj.reset_code:
                                obj.reset_text = translate_condition(obj.reset_code, xml_cache)
                                
                        add_log(f"✍️ Обновление текстовых описаний в конфигураторе...", "INFO")
                        update_configurator_texts(fp, parsed_objects, config_class, add_log)
                    # ------------------------------------------------

                    # --- БЛОК ОБНОВЛЕНИЯ ДОКУМЕНТА (xlwings) ---
                    if (is_ok or force) and do_document and getattr(config_class, 'SUPPORT_DOC_UPDATE', False):
                        add_log(f"Обновление документа {clean_n}...", "INFO")
                        parsed_objects = getattr(parser, 'all_parsed_objects', [])
                        update_configurator_document(fp, parsed_objects, config_class, add_log)
                    # ------------------------------------------------

                    # --- БЛОК ПОКРАСКИ ЯЧЕЕК (xlwings) ---
                    if (is_ok or force) and do_coloring and getattr(config_class, 'SUPPORT_COLOR_UPDATE', False):
                        add_log(f"🎨 Анализ и обновление цветов ячеек...", "INFO")
                        # Передаем карту покраски, которую сформировал парсер
                        rows_to_color = getattr(parser, 'rows_to_color', {})
                        update_configurator_colors(fp, rows_to_color, config_class, add_log)
                    # ------------------------------------------------

                    st.session_state.file_errors.extend(parser.errors)
                    if (is_ok or force) and do_sources:
                        st.session_state.files_to_write.extend(parser.files_to_write)

            if st.session_state.file_errors and not force:
                st.session_state.process_step = "awaiting_confirm"
                st.rerun()
            else:
                if st.session_state.files_to_write:
                    for item in st.session_state.files_to_write:
                        os.makedirs(os.path.dirname(item['path']), exist_ok=True)
                        with open(item['path'], "w", encoding=AppConfig.FILE_ENCODING) as f: 
                            f.write(item['text'])
                    add_log(f"✨ Процесс завершен! Файлов создано: {len(st.session_state.files_to_write)}")
                st.session_state.process_step = "done"

    # Вызов нашей вынесенной функции для отрисовки логов
    render_logs()

    from application.logger import render_xml_inspector
    render_xml_inspector()