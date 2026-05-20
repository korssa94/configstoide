"""Перекрестная проверка сигналов конфигуратора Сигнализаций (ТБ51)
против связанных конфигураторов Входов/Выходов (ТЭ5).

Использует Мастер-конфигуратор для поиска связей. Читает каждый связанный
ТЭ5 ровно один раз, формирует кэш дефектных строк для последующего
force-режима (повторного запуска "всё равно создать").

Все хелперы с `_` — реализация модуля. Публичный entry point — cross_validate_alarms.
"""
import os
import openpyxl
import streamlit as st

from inout_configurator.config import TE5Config
from application.settings.app_config import AppConfig


# --- Публичная точка входа ---
def cross_validate_alarms(parser, filepath, clean_name, force=False):
    """Главная точка входа кросс-проверки.

    Возвращает True, если кросс-проверка пройдена или пропущена (мастер не найден,
    связей нет и т.п.). False — если найдены ошибки и пользователь должен принять решение.
    """
    # FORCE-режим: используем кэш дефектных строк, собранный на этапе анализа
    if force:
        cached = _get_cached_failed_rows(clean_name)
        if cached is not None:
            if cached:
                _exclude_failed_rows(parser, cached)
            return True
        # Кэша нет (свежий процесс или сбой анализа) — проваливаемся в полный прогон

    # ANALYZE-режим: полный прогон
    parser.log(f"🚀 Старт кросс-проверки (сбор данных) для {clean_name}", level="INFO")

    linked_inout_files = _read_inout_links_from_master(parser, clean_name, filepath)
    if linked_inout_files is None:
        return False  # упало чтение мастера — это критическая ошибка
    if not linked_inout_files:
        return True   # связи не найдены — проверка пропускается (warning уже в логе)

    inout_groups, ctrl_to_inout_file = _group_controllers_by_inout(
        parser, linked_inout_files, base_dir=os.path.dirname(filepath)
    )
    inout_signals = _read_inout_signals(parser, inout_groups)
    parser.log("⏱️ Сбор данных для проверки завершен", level="INFO")

    failed_rows = _compare_alarms_with_inout(
        parser, inout_signals, ctrl_to_inout_file, clean_name, force=False
    )
    _cache_failed_rows(clean_name, failed_rows)

    if failed_rows:
        parser.log(
            f"❌ Перекрестная проверка завершена. Найдено ошибочных строк: {len(failed_rows)}",
            level="ERROR",
        )
        return False

    parser.log("✅ Перекрестная валидация ТБ51 и ТЭ5 успешно пройдена для всех контроллеров!", level="INFO")
    return True


# --- Кэш дефектных строк (Streamlit session_state) ---
def _get_cached_failed_rows(clean_name):
    """Возвращает set дефектных строк из кэша. None — если кэш ещё не инициализирован."""
    if 'failed_rows_cache' not in st.session_state:
        return None
    return st.session_state.failed_rows_cache.get(clean_name, set())


def _cache_failed_rows(clean_name, failed_rows):
    """Сохраняет результат проверки в st.session_state для возможного force-прогона."""
    if 'failed_rows_cache' not in st.session_state:
        st.session_state.failed_rows_cache = {}
    st.session_state.failed_rows_cache[clean_name] = failed_rows


def _exclude_failed_rows(parser, failed_rows):
    """В force-режиме: пишет WARNING по каждой дефектной строке и вырезает её из self.objects."""
    ws_alarms = parser.wb_data[parser.config.SHEET_ALARMS]
    col_map = parser.get_column_mapping(ws_alarms, parser.config.HEADER_ROW)
    col_param_code = getattr(parser.config, "COL_PARAM_CODE", "Алг.имя сигнала")

    for r in sorted(list(failed_rows)):
        p_name = _get_cell_val(parser, ws_alarms, col_map, r, col_param_code)
        p_desc = _get_cell_val(parser, ws_alarms, col_map, r, parser.config.COL_MESSAGE)
        parser.log(
            f"⚠️ Переменная со строки {r}, наименованием '{p_desc or '---'}', "
            f"алг. именем '{p_name or '---'}' исключена из генерации (провал кросс-проверки)",
            level="WARNING",
        )

    def get_obj_row(o):
        return getattr(o, 'row_number', getattr(o, 'row', None))

    parser.all_parsed_objects = [obj for obj in parser.all_parsed_objects if get_obj_row(obj) not in failed_rows]
    for prefix in parser.objects.keys():
        parser.objects[prefix] = [obj for obj in parser.objects[prefix] if get_obj_row(obj) not in failed_rows]


# --- Чтение Мастер-конфигуратора ---
def _read_inout_links_from_master(parser, clean_name, filepath):
    """Открывает Мастер-конфигуратор, ищет ключевые строки (Контроллер/Входы-Выходы/Сигнализации).
    Возвращает список [{controller, inout_file}, ...].

    Различия в возвратах:
      None  — упало с исключением (критическая ошибка)
      []    — мастера нет / ключевых строк нет / связей нет (пропускаем проверку, warning в лог)
    """
    master_path = getattr(parser, 'master_file', None)
    if not master_path or not os.path.exists(master_path):
        parser.log("⚠️ Путь к Мастер-конфигуратору не передан или файл не найден. Кросс-проверка пропущена.", level="WARNING")
        return []

    parser.log("📂 Чтение карты связей из Мастер-конфигуратора...", level="INFO")
    alarm_filename = os.path.basename(filepath)
    linked = []

    try:
        wb_master = openpyxl.load_workbook(master_path, data_only=True, read_only=True)
        ws_master = wb_master.active

        rows = list(ws_master.iter_rows(min_row=1, max_row=50, values_only=True))

        row_ctrl  = next((i for i, r in enumerate(rows) if r[0] and AppConfig.MASTER_ROW_CTRL  in str(r[0]).lower()), None)
        row_inout = next((i for i, r in enumerate(rows) if r[0] and AppConfig.MASTER_ROW_TE5   in str(r[0]).lower()), None)
        row_alarm = next((i for i, r in enumerate(rows) if r[0] and AppConfig.MASTER_ROW_TB51  in str(r[0]).lower()), None)

        if None in (row_ctrl, row_inout, row_alarm):
            parser.log("⚠️ Не найдены ключевые строки в Мастер-конфигураторе.", "WARNING")
            wb_master.close()
            return []

        for col_idx in range(1, len(rows[row_ctrl])):
            alarm_val = rows[row_alarm][col_idx]
            if alarm_val and (clean_name in str(alarm_val).strip() or str(alarm_val).strip() == alarm_filename):
                ctrl = rows[row_ctrl][col_idx]
                inout = rows[row_inout][col_idx]
                if ctrl and inout:
                    linked.append({
                        "controller": str(ctrl).strip(),
                        "inout_file": str(inout).strip(),
                    })

        wb_master.close()
        parser.log(
            f"🔗 Найдено связей: {len(linked)}. Контроллеры: {[link['controller'] for link in linked]}",
            level="INFO",
        )
    except Exception as e:
        parser.log(f"❌ Ошибка при чтении Мастер-конфигуратора: {e}", level="ERROR")
        return None

    if not linked:
        parser.log("⚠️ В Мастер-конфигураторе не найдено привязанных ТЭ5 для этого ТБ51.", level="WARNING")

    return linked


# --- Группировка контроллеров по уникальным путям ТЭ5 ---
def _group_controllers_by_inout(parser, linked, base_dir):
    """Принимает [{controller, inout_file}, ...]. Ищет реальные пути ТЭ5 в base_dir
    (имя в мастере — это префикс), группирует контроллеры по уникальным путям.
    Возвращает (inout_groups, ctrl_to_inout_file).
    """
    inout_groups = {}        # inout_path -> [список контроллеров]
    ctrl_to_inout_file = {}  # ctrl_name -> имя файла ТЭ5 (basename)

    for link in linked:
        ctrl_name = link['controller']
        inout_base_name = link['inout_file']
        inout_path = None

        if os.path.exists(base_dir):
            for f in os.listdir(base_dir):
                if f.startswith(inout_base_name) and f.endswith(('.xlsm', '.xlsx')) and not f.startswith('~$'):
                    inout_path = os.path.join(base_dir, f)
                    break

        if not inout_path:
            parser.log(
                f"⚠️ Файл для {inout_base_name} (контроллер {ctrl_name}) не найден в папке. Пропуск.",
                level="WARNING",
            )
            continue

        ctrl_to_inout_file[ctrl_name] = os.path.basename(inout_path)
        inout_groups.setdefault(inout_path, []).append(ctrl_name)

    return inout_groups, ctrl_to_inout_file


# --- Чтение сигналов из связанных ТЭ5 ---
def _read_inout_signals(parser, inout_groups):
    """Для каждого уникального пути ТЭ5 — открывает файл ровно один раз, извлекает
    сигналы из всех листов MODEL_SETTINGS, раскидывает по контроллерам группы.
    Возвращает {ctrl_name: [(sig_type, alg_name, active_setpoints), ...]}.
    """
    setpoint_header_map = _build_setpoint_header_map()
    inout_signals = {}

    for inout_path, ctrls in inout_groups.items():
        fname = os.path.basename(inout_path)
        ctrls_str = ", ".join(ctrls)
        parser.log(f"📂 Быстрое чтение {fname} для ПЛК {ctrls_str}...", level="INFO")

        extracted = []
        try:
            wb_inout = openpyxl.load_workbook(inout_path, data_only=True, read_only=True)

            for sig_type, settings in TE5Config.MODEL_SETTINGS.items():
                sheet_name = settings.get("sheet_name")
                if sheet_name not in wb_inout.sheetnames:
                    continue

                ws_inout = wb_inout[sheet_name]
                alg_col_idx, create_col_idx, setpoint_cols = _find_inout_header_columns(
                    ws_inout, setpoint_header_map
                )

                if alg_col_idx is None or create_col_idx is None:
                    continue

                for row in ws_inout.iter_rows(min_row=TE5Config.DATA_START_ROW, values_only=True):
                    max_idx = max([alg_col_idx, create_col_idx] + list(setpoint_cols.values()))
                    if len(row) <= max_idx:
                        continue

                    alg_val = row[alg_col_idx]
                    create_val = row[create_col_idx]

                    if not alg_val or not str(alg_val).strip():
                        continue
                    if str(create_val).strip() != "1" and create_val != 1:
                        continue

                    active_setpoints = []
                    if sig_type.lower() == "taipar":
                        for sp_name, sp_col in setpoint_cols.items():
                            sp_val = row[sp_col]
                            if sp_val is not None and str(sp_val).strip() != "":
                                active_setpoints.append(sp_name)

                    extracted.append((sig_type.lower(), str(alg_val).strip(), active_setpoints))

            wb_inout.close()

            for ctrl_name in ctrls:
                inout_signals[ctrl_name] = extracted

        except Exception as e:
            parser.log(f"❌ Ошибка при чтении {fname}: {e}", level="ERROR")

    return inout_signals


def _find_inout_header_columns(ws_inout, setpoint_header_map):
    """Ищет в HEADER_ROW индексы колонок: алг.имя, "Создавать код", уставки.
    Возвращает (alg_col_idx, create_col_idx, {sp_name: col_idx}).
    """
    alg_col_idx = None
    create_col_idx = None
    setpoint_cols = {}

    header_row_data = next(
        ws_inout.iter_rows(min_row=TE5Config.HEADER_ROW, max_row=TE5Config.HEADER_ROW, values_only=True),
        None,
    )
    if header_row_data is None:
        return alg_col_idx, create_col_idx, setpoint_cols

    for col_idx, cell_val in enumerate(header_row_data):
        if not cell_val:
            continue
        header = str(cell_val).strip().lower()
        if header == TE5Config.COL_ALG_NAME.lower():
            alg_col_idx = col_idx
        elif header == TE5Config.COL_CREATE_CODE.lower():
            create_col_idx = col_idx
        elif header in setpoint_header_map:
            setpoint_cols[setpoint_header_map[header]] = col_idx

    return alg_col_idx, create_col_idx, setpoint_cols


def _build_setpoint_header_map():
    """Карта {русский_заголовок_lower: латинский_код_уставки}."""
    return {
        TE5Config.COL_LL.lower().strip(): "ll",
        TE5Config.COL_L1.lower().strip(): "l1",
        TE5Config.COL_L.lower().strip():  "l",
        TE5Config.COL_H.lower().strip():  "h",
        TE5Config.COL_H1.lower().strip(): "h1",
        TE5Config.COL_HH.lower().strip(): "hh",
    }


# --- Сравнение строк ТБ51 с собранными сигналами ТЭ5 ---
def _compare_alarms_with_inout(parser, inout_signals, ctrl_to_inout_file, clean_name, force):
    """Идёт по строкам листа Сигнализаций, для каждой строки с пустым "Условие (код)"
    проверяет наличие соответствующего параметра в собранных сигналах ТЭ5.
    Возвращает set дефектных строк.
    """
    parser.log("🔍 Запуск перекрестной проверки сигналов...", level="INFO")

    # Хэш-таблица для O(1)-поиска: {ctrl: {(type, alg_name_lower): set_of_active_setpoints}}
    inout_lookup = {}
    for ctrl, signals in inout_signals.items():
        inout_lookup[ctrl] = {}
        for item in signals:
            s_type = item[0].lower()
            s_name = item[1].lower()
            s_sps = set(item[2]) if len(item) > 2 else set()
            inout_lookup[ctrl][(s_type, s_name)] = s_sps

    ws_alarms = parser.wb_data[parser.config.SHEET_ALARMS]
    col_map = parser.get_column_mapping(ws_alarms, parser.config.HEADER_ROW)

    col_cond_code  = getattr(parser.config, "COL_CONDITION_CODE", "Условие (код)")
    col_setpoint   = getattr(parser.config, "COL_SETPOINT",       "Уставка")
    col_condition  = getattr(parser.config, "COL_CONDITION",      "Условие")
    col_param_code = getattr(parser.config, "COL_PARAM_CODE",     "Алг.имя сигнала")

    failed_rows = set()

    for r in range(parser.config.DATA_START_ROW, ws_alarms.max_row + 1):
        cond_code = _get_cell_val(parser, ws_alarms, col_map, r, col_cond_code)
        if cond_code != "":
            continue  # код прописан вручную — кросс-чек неприменим

        setpoint   = _get_cell_val(parser, ws_alarms, col_map, r, col_setpoint)
        condition  = _get_cell_val(parser, ws_alarms, col_map, r, col_condition)
        param_code = _get_cell_val(parser, ws_alarms, col_map, r, col_param_code)

        if not param_code or "резерв" in param_code.lower():
            continue

        expected_type = _detect_expected_type(setpoint, condition)
        if not expected_type:
            continue

        missing_param_ctrls, missing_sp_ctrls = _find_missing_in_inout(
            inout_lookup, inout_signals.keys(), expected_type, param_code, setpoint
        )

        if missing_param_ctrls:
            _report_missing_param(
                parser, r, param_code, expected_type, missing_param_ctrls,
                ctrl_to_inout_file, force, failed_rows
            )
        if missing_sp_ctrls:
            _report_missing_setpoint(
                parser, r, param_code, setpoint, missing_sp_ctrls,
                ctrl_to_inout_file, force, failed_rows
            )

    return failed_rows


def _detect_expected_type(setpoint, condition):
    """По колонкам "Уставка" и "Условие" определяет, какой тип параметра ожидается в ТЭ5:
    'tdipar' (дискретный), 'taipar' (аналоговый), None (не проверяем).
    """
    if (setpoint == "" and condition == "DI") or setpoint == "N":
        return "tdipar"
    if setpoint in ["LL", "L1", "L", "H", "H1", "HH"]:
        return "taipar"
    return None


def _inout_sheet_for_type(expected_type):
    """Возвращает имя листа в ТЭ5 для ожидаемого типа параметра.
    Пример: 'tdipar' → 'Вх.Д сигн.', 'taipar' → 'Вх.А сигн.'

    Так в сообщения об ошибках можно положить понятное "где смотреть", а не
    внутренний код типа.
    """
    for name, settings in TE5Config.MODEL_SETTINGS.items():
        if name.lower() == expected_type.lower():
            return settings.get("sheet_name", expected_type)
    return expected_type  # фоллбэк на случай неизвестного типа


def _find_missing_in_inout(inout_lookup, ctrls, expected_type, param_code, setpoint):
    """Для каждого контроллера: проверяет наличие параметра и активность нужной уставки.
    Возвращает (missing_param_ctrls, missing_sp_ctrls).
    """
    missing_param = []
    missing_sp = []

    for ctrl in ctrls:
        lookup_dict = inout_lookup.get(ctrl, {})
        key = (expected_type, param_code.lower())

        if key not in lookup_dict:
            missing_param.append(ctrl)
            continue

        # Параметр есть. Если аналоговый — проверяем, что нужная уставка активна.
        if expected_type == "taipar":
            sp_to_check = setpoint.lower()
            if sp_to_check in ["ll", "l1", "l", "h", "h1", "hh"]:
                active_sps = lookup_dict[key]
                if sp_to_check not in active_sps:
                    missing_sp.append(ctrl)

    return missing_param, missing_sp


def _report_missing_param(parser, row, param_code, expected_type, missing_ctrls,
                          ctrl_to_inout_file, force, failed_rows):
    """Группирует "отсутствующие" контроллеры по файлам ТЭ5 и пишет ERROR по каждой группе."""
    sheet_name = _inout_sheet_for_type(expected_type)
    for fname, ctrls in _group_ctrls_by_file(missing_ctrls, ctrl_to_inout_file).items():
        failed_rows.add(row)
        if not force:
            err_msg = (
                f"Строка {row}: [{os.path.splitext(fname)[0]}, ПЛК {', '.join(ctrls)}] "
                f"Сигнал '{param_code}' не найден в ТЭ5 на листе '{sheet_name}'."
            )
            parser.log(err_msg, level="ERROR")
            parser.errors.append(err_msg)


def _report_missing_setpoint(parser, row, param_code, setpoint, missing_ctrls,
                             ctrl_to_inout_file, force, failed_rows):
    """Аналогично _report_missing_param, но для "параметр есть, уставка пустая"."""
    for fname, ctrls in _group_ctrls_by_file(missing_ctrls, ctrl_to_inout_file).items():
        failed_rows.add(row)
        if not force:
            err_msg = (
                f"Строка {row}: [{os.path.splitext(fname)[0]}, ПЛК {', '.join(ctrls)}] "
                f"У параметра '{param_code}' в ТЭ5 не задана (пустая) уставка '{setpoint}'."
            )
            parser.log(err_msg, level="ERROR")
            parser.errors.append(err_msg)


def _group_ctrls_by_file(ctrls, ctrl_to_inout_file):
    """Группирует контроллеры по файлам ТЭ5, к которым они привязаны.
    Возвращает {filename: [ctrl1, ctrl2, ...]}.
    """
    groups = {}
    for ctrl in ctrls:
        fname = ctrl_to_inout_file.get(ctrl, "Неизвестный ТЭ5")
        groups.setdefault(fname, []).append(ctrl)
    return groups


# --- Низкоуровневый хелпер чтения ячейки ---
def _get_cell_val(parser, ws, col_map, row_idx, col_name):
    """Безопасное чтение ячейки по имени колонки. Возвращает strip()'нутую строку или ""."""
    idx = parser.find_col_idx(col_map, col_name)
    val = ws.cell(row=row_idx, column=idx).value if idx else ""
    return str(val).strip() if val is not None else ""