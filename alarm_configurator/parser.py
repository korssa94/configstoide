import openpyxl
import os
import re
import datetime
from shared.parsers.base_parser import BaseParser
from alarm_configurator.models.talr import Talr
from alarm_configurator.models.tppu import Tppu
from alarm_configurator.models.tcrs import Tcrs
from alarm_configurator.cross_validation import cross_validate_alarms

class AlarmParser(BaseParser):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def parse(self, clean_name, active_rules, te5_objects=None, skip_paint=False, force=False):
        try:
            wb_data = openpyxl.load_workbook(self.filepath, data_only=True)
            self.file_author = wb_data.properties.lastModifiedBy or "Unknown"
            if wb_data.properties.modified:
                utc_time = wb_data.properties.modified
                local_time = utc_time.replace(tzinfo=datetime.timezone.utc).astimezone(None)
                self.file_save_time = local_time.strftime("%Y-%m-%d %H:%M:%S")
            
            algo_folder = self.config.DEFAULT_ALGO_FOLDER

            # --- 1. СБОР ПОДСИСТЕМ (в методе parse) ---
            subsystems_map = {}
            if self.config.SHEET_SETTINGS in wb_data.sheetnames:
                ws_settings = wb_data[self.config.SHEET_SETTINGS]
                # Предположим, что структура: Название(0), Код(1), Номер(2)
                # Начинаем читать после заголовка "Подсистемы"
                start_reading = False
                for row in ws_settings.iter_rows(values_only=True):
                    if not row or not row[0]: continue
                    if "Подсистемы" in str(row[0]):
                        start_reading = True
                        continue
                    if start_reading:
                        name = str(row[0]).strip()
                        code = str(row[1]).strip()
                        num = str(row[2]).strip()
                        if code:
                            subsystems_map[code] = {"num": num, "name": name}
                            
                self.subsystems_map = subsystems_map # Запоминаем подсистемы в классе для шага генерации GVL

            # --- 2. ПАРСИНГ СИГНАЛИЗАЦИЙ ---
            all_parsed_objects = []
            current_sub_code = "" # Переменная-память

            if self.config.SHEET_ALARMS not in wb_data.sheetnames:
                self.log(f'Лист "{self.config.SHEET_ALARMS}" не найден.', level="ERROR")
                return False
            
            ws_data = wb_data[self.config.SHEET_ALARMS]
            col_map = self.get_column_mapping(ws_data, self.config.HEADER_ROW)

            # Подготавливаем словари для объектов (alr, trs, lmt и т.д.)
            for key in self.config.MODEL_SETTINGS.keys():
                self.objects[key] = []

            for row in range(self.config.DATA_START_ROW, ws_data.max_row + 1):
                def get_val(col_name):
                    idx = self.find_col_idx(col_map, col_name)
                    val = ws_data.cell(row=row, column=idx).value if idx else ""
                    return val if val is not None else ""
                
                type_str = str(get_val(self.config.COL_TYPE)).strip()
                msg = str(get_val(self.config.COL_MESSAGE)).strip()
                action = str(get_val(self.config.COL_ACTION)).strip()

                # ПРОВЕРКА: Это строка подсистемы? (VBA: Сообщение <> "" And Тип = "")
                if msg != "" and type_str == "":
                    # Алгоритмическое имя подсистемы в этой строке
                    current_sub_code = str(get_val(self.config.COL_ALG_NAME)).strip()
                    continue

                # 1. Анализируем строку (как в VBA)
                target_prefixes = []

                if type_str == self.config.ALARM_TYPE_WARNING:
                    if action in self.config.ACTIONS_ADDING_PPU:
                        target_prefixes = ["alr", "ppu"]
                    else:
                        target_prefixes = ["alr"]

                elif type_str == self.config.ALARM_TYPE_PRESTART:
                    target_prefixes = ["ppu"]

                elif type_str == self.config.ALARM_TYPE_EMERGENCY:
                    if action in self.config.ACTIONS_MAKING_CRS:
                        target_prefixes = ["crs"]  # Аварийная
                    else:
                        target_prefixes = ["trs"]  # Тревожная

                elif type_str == self.config.ALARM_TYPE_LIMITING:
                    target_prefixes = ["lmt"]  # Ограничительная

                # 2. Генерируем объекты по списку целей
                for prefix in target_prefixes:
                    if prefix in ["alr", "trs", "lmt"]:
                        obj = Talr(row, get_val, self.config, prefix=prefix)
                    elif prefix == "ppu":
                        obj = Tppu(row, get_val, self.config, prefix=prefix)
                    elif prefix == "crs":
                        obj = Tcrs(row, get_val, self.config, prefix=prefix)
                    else:
                        continue
                    
                    # Присваиваем данные подсистемы из памяти
                    sub_info = subsystems_map.get(current_sub_code, {})
                    obj.subsystem_code = current_sub_code
                    obj.subsystem_num = sub_info.get("num", "")
                    obj.subsystem_name = sub_info.get("name", "")

                    # Раскладываем по нужным спискам
                    all_parsed_objects.append(obj)
                    if obj.prefix in self.objects:
                        self.objects[obj.prefix].append(obj)

            # Сохраняем книгу и плоский список объектов внутри класса
            self.wb_data = wb_data 
            self.all_parsed_objects = all_parsed_objects

            # Запоминаем начальное количество сигналов по типам для статистики
            initial_counts = {prefix: len(items) for prefix, items in self.objects.items()}

            # --- 1. ПЕРВИЧНЫЙ АНАЛИЗ (Выводим ТОЛЬКО при первом запуске) ---
            if not force:
                for prefix, initial_cnt in initial_counts.items():
                    if initial_cnt > 0:
                        self.log(f'Собрано {initial_cnt} сигналов типа "{prefix}".', level="INFO")

            # --- 2. КРОСС-ПРОВЕРКА С ТЭ5 ---
            is_valid = cross_validate_alarms(self, self.filepath, clean_name, force=force)
            
            if not is_valid and not force:
                self.log("❌ Обнаружены ошибки кросс-проверки. Процесс остановлен.", level="ERROR")
                return False

            # --- 3. ФИНАЛЬНЫЕ ИТОГИ ГЕНЕРАЦИИ (Выводим ТОЛЬКО при force=True) ---
            if force:
                for prefix, initial_cnt in initial_counts.items():
                    current_cnt = len(self.objects.get(prefix, []))
                    skipped_cnt = initial_cnt - current_cnt
                    
                    if skipped_cnt > 0:
                        self.log(f'Тип "{prefix}": создано {current_cnt} сигналов, пропущено {skipped_cnt} из-за ошибок кросс-проверки.', level="INFO")
                    else:
                        if current_cnt > 0:
                            self.log(f'Тип "{prefix}": создано {current_cnt} сигналов.', level="INFO")

            # --- 4. ГЕНЕРАЦИЯ ФАЙЛОВ ---
            self._generate_files(algo_folder, self.all_parsed_objects)
            
            return True

        except Exception as e:
            self.log(f"Ошибка в парсере Сигнализаций: {str(e)}", level="ERROR")
            return False

    def _generate_files(self, algo_folder, all_parsed_objects):
        fn = os.path.basename(self.filepath)
        excel_v = (re.search(r'v\d+\.\d+\.\d+', fn).group(0) if re.search(r'v\d+\.\d+\.\d+', fn) else "none")
        
        # ШАГ 1: ГЕНЕРАЦИЯ GVL (Индивидуально для каждого префикса)
        for prefix, settings in self.config.MODEL_SETTINGS.items():
            prefix_objects = self.objects.get(prefix, [])
            
            # 1. Системная шапка (создается всегда)
            gvl_txt = self.generate_header(algo_folder, self.file_author, self.file_save_time, excel_v, settings.get("desc_gvl", ""))
            
            # 2. Декларативная шапка ТИПА
            if prefix_objects:
                # Если объекты есть, берем специфичную шапку у первого (например, для ППУ)
                gvl_txt += prefix_objects[0].get_gvl_header()
            else:
                # Если объектов нет, выводим базовую структуру (как в твоем примере ОС)
                gvl_txt += "{attribute 'qualified_only'}\n"
                gvl_txt += "{attribute 'symbol' := 'read'}\n"
                gvl_txt += "VAR_GLOBAL\n"
                gvl_txt += "\t{attribute 'symbol' := 'none'}\n"
                gvl_txt += "\tcommon : Talarm; //Общий сигнал для VAR_STAT\n\n"

            # 3. Сами сигналы (только если они есть)
            if prefix_objects:
                gvl_txt += "\n".join([o.get_gvl_string() for o in prefix_objects]) + "\n"
            
            gvl_txt += "END_VAR\n"
            
            for ctrl in self.matched_ctrls:
                tdir = os.path.join(self.base_dir, self.config.SOURCE_FOLDER, ctrl, algo_folder)
                self.files_to_write.append({"path": os.path.join(tdir, settings["file_gvl"]), "text": gvl_txt})


        # ШАГ 2: ГЕНЕРАЦИЯ ST (Собираем все префиксы в общие ST-файлы по порядку)
        st_grouped = {}
        
        # all_parsed_objects - это список всех объектов строго в порядке строк из Excel
        for obj in all_parsed_objects: 
            # Узнаем, в какой ST-файл должен идти этот объект (например, alr_update.st)
            settings = self.config.MODEL_SETTINGS.get(obj.prefix, {})
            f_st = settings.get("file_st")
            
            if not f_st:
                continue # Если файл ST не указан, пропускаем
                
            # Инициализируем хранилище для файла, если его еще нет
            if f_st not in st_grouped:
                st_grouped[f_st] = {
                    "desc": settings.get("desc_st", ""), 
                    "global_var": settings.get("global_var", "alr"), 
                    "content": "",
                    "last_subsystem": None # Память для заголовков подсистем ВНУТРИ файла
                }
            elif settings.get("desc_st"):
                st_grouped[f_st]["desc"] = settings["desc_st"]
                st_grouped[f_st]["global_var"] = settings["global_var"]

            # Логика отрисовки красивых заголовков подсистем
            if obj.subsystem_name != st_grouped[f_st]["last_subsystem"]:
                if st_grouped[f_st]["content"] == "":
                        st_grouped[f_st]["content"] += "//==========================================================================\n\n"
                st_grouped[f_st]["content"] += "//==========================================================================\n"
                st_grouped[f_st]["content"] += f"//Подсистема: {obj.subsystem_name}\n"
                st_grouped[f_st]["content"] += "//==========================================================================\n"
                st_grouped[f_st]["last_subsystem"] = obj.subsystem_name
            
            # Добавляем вызов самого блока
            st_grouped[f_st]["content"] += obj.get_st_string() + "\n"

        # ШАГ 3: Сохраняем сгруппированные ST файлы
        for f_st, data in st_grouped.items():
            h_st = self.generate_header(algo_folder, self.file_author, self.file_save_time, excel_v, data["desc"])
            st_txt = h_st + f"FUNCTION {data['global_var']}_update : BOOL\nVAR_INPUT\nEND_VAR\n"
            st_txt += "//++++++< Content:Implementation >++++++++++++++++++++++++++++++++++++++++++++//\n"
            st_txt += data["content"]
            
            for ctrl in self.matched_ctrls:
                tdir = os.path.join(self.base_dir, self.config.SOURCE_FOLDER, ctrl, algo_folder)
                self.files_to_write.append({"path": os.path.join(tdir, f_st), "text": st_txt})

        # --- ШАГ 4: ГЕНЕРАЦИЯ ФАЙЛА СБРОСА reset_alarm_var_stat.st ---
        reset_txt = (
            f"//++++++< Content:Path >++++++++++++++++++++++++++++++++++++++++++++++++++++++//\n"
            f"//{algo_folder}\n"
            f"//++++++< Content:Declaration >+++++++++++++++++++++++++++++++++++++++++++++++//\n"
            f"{{region INFO}}\n"
            f"(*\n"
            f" Назначение: Используется для сброса всех VAR_STAT сигнализаций\n"
            f" Автор:   /{self.file_author}/ {self.file_save_time}\n"
            f" Версия:  {excel_v}\n"
            f" Макросы: v1.0.0\n"
            f"*)\n"
            f"{{endregion}}\n"
            f"FUNCTION reset_alarm_var_stat : BOOL\n"
            f"VAR_INPUT\n"
            f"END_VAR\n"
            f"//++++++< Content:Implementation >++++++++++++++++++++++++++++++++++++++++++++//\n"
            f"//Сбрасываем обобщенные сигнализации\n"
            f"//По типу и подсистемам в 0, потому что при сработке они выставляются в 1\n"
            f"crs.common.type_aggregated := crs.common.subsystem_aggregated := 0;\n"
            f"//По взводу защит все биты в 1, потому что при отсутствии взвода они сбрасываются в 0\n"
            f"crs.common.arm_aggregated := 16#FFFFFFFF;\n"
            f"//По типу биты 1-3 в 0 (для alr, lmt и trs, потому что при сработке они выставляются в 1), 4 и 5 в 1 (для ppu, потому что при сработке они сбрасываются в 0)\n"
            f"alr.common.type_aggregated := 2#1110000;\n"
            f"//Для типов alr, lmt и trs по подсистемам все биты в 0, потому что при сработке они выставляются в 1\n"
            f"alr.common.subsystem_aggregated[1] := alr.common.subsystem_aggregated[2] := alr.common.subsystem_aggregated[3] := 0;\n"
            f"//Для ППУ по подсистемам все биты в 1, потому что при сработке они сбрасываются в 0\n"
            f"alr.common.subsystem_aggregated[4] := alr.common.subsystem_aggregated[5] := 16#FFFFFFFF;\n"
        )
        
        for ctrl in self.matched_ctrls:
            tdir = os.path.join(self.base_dir, self.config.SOURCE_FOLDER, ctrl, algo_folder)
            self.files_to_write.append({"path": os.path.join(tdir, "reset_alarm_var_stat.st"), "text": reset_txt})

        # --- ШАГ 5: ГЕНЕРАЦИЯ subsystem.gvl ДЛЯ ВСЕХ КОНТРОЛЛЕРОВ ---
        sub_list = [f"{info['name']} - {code} - {info['num']}" for code, info in getattr(self, 'subsystems_map', {}).items()]
        self.generate_subsystems_gvl(sub_list, "сигнализаций")
            
        return True
  
